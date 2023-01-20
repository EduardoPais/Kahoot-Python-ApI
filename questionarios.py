import openpyxl
import os
import time


def criar():
    from autenticacao import email_log
    book = openpyxl.load_workbook("questionarios.xlsx")
    z=1
    c=1
    a=0
    b=1
    f=0
    g=0
    #cria nova pagina
    while g!=1:
        os.system("cls")
        print("\n*******Modo Criação*******")
        nome=input("\nDigite nome para o teste: ")
        if nome in book.sheetnames:
            print("Já existe um teste com este nome!")
        elif nome.lower() =="sair":
            print("Nome não premitido!")
        else:
            book.create_sheet(nome)
            page = book[nome]
            
            #escreve nome do criador para que este não posso responder ao proprio questionário
            book2 = openpyxl.load_workbook("respostas.xlsx")
            book2.create_sheet(nome)
            page2 = book2[nome]
            cell5 = page2.cell(row=1,column=1)
            cell5.value=email_log
            cell5 = page2.cell(row=1,column=2)
            cell5.value="criador"
            g=1
    #valida se as perguntas nao passam de 10
    while a!=1:
        os.system("cls")
        VF=int(input("\nDigite o numero de perguntas (maximo 10) de verdadeiro ou falso que quer: "))
        EM=int(input("\nDigite o numero de perguntas (maximo 10) de escolha multipla que quer: "))
        if VF==0 and EM==0:
            print("Tem de ter pelo menos uma pergunta!")
        elif 0<=VF<=10 and 0<=EM<=10:
            a=1
    cell2=page.cell(row=z,column=1)
    cell3=page.cell(row=z,column=2)
    if VF!=0:
        os.system("cls")
        print("\nPerguntas de Verdadeiro ou Falso:\n")
    #escrever perguntas e suas respostas
    while z<=VF:
        cell2=page.cell(row=z,column=1)
        cell3=page.cell(row=z,column=2)
        print(b,"º-",end="")
        cell2.value=input("Pergunta: ")
        res=input("\nResposta: ")
        #validaçao das respostas
        f=0
        if res.lower() != "v" and res.lower() != "f" and res.lower() != "verdadeiro" and res.lower() != "falso" and cell2.value==None:
            print("Resposta invalida!\n")
            z-=1
            b-=1
            f=1
        if res.lower() == "v" or res.lower() == "verdadeiro":
            cell3.value="v"
        if res.lower() == "f" or res.lower() == "falso":
            cell3.value="f"
        #modo de rescrever as perguntas
        while f !=1:
            cont=input("\nEstá todo bem escrito:\nSim-continuas\nNão-refazer\n")
            if cont.lower()=="não" or cont.lower()=="nao":
                z-=1
                b-=1
                f=1
            elif cont.lower() == "sim":
                f=1
            else:
                print("\nRespoata invalida!")

        z+=1
        b+=1
    if EM!=0:
        os.system("cls")
        print("\nPerguntas de Escolha Multipla:\n")
    b=1
    #escrever perguntas e suas respostas
    while c<=EM:
        cell2=page.cell(row=z+c,column=1)
        cell3=page.cell(row=z+c,column=2)
        print(b,"º-",end="")
        cell2.value=input("Pergunta: ")
        cell3.value=input("\nResposta: ")
        #validaçao das respostas
        d=1
        if cell3.value.lower() != "a" and cell3.value.lower() != "b" and cell3.value.lower() != "c" and cell3.value.lower() != "d":
            print("\nResposta invalida!\n")
            c-=1
            b-=1
            d=6
        e=3
        #escrever informaçao de cada alinia
        while d<=4:
            cell4=page.cell(row=z+c,column=e)
            if d==1:
                em="A- "
            if d==2:
                em="B- "
            if d==3:
                em="C- "
            if d==4:
                em="D- "
            cell4.value=input(em)
            d+=1
            e+=1
        #modo de rescrever as perguntas
        while d!=6:
            cont=input("\nEstá todo bem escrito:\nSim-continuas\nNão-refazer\n")    
            if cont.lower()=="não" or cont.lower()=="nao":
                c-=1
                b-=1
                d=6
            elif cont.lower() == "sim":
                d=6
            else:
                print("\nRespoata invalida!")
        c+=1
        b+=1

    print("Nome do questionário:",nome)
    time.sleep(3)
    book.save("questionarios.xlsx")
    book2.save("respostas.xlsx")

def responder():
    from autenticacao import email_log
    a=0
    b=1
    x=0
    c=1
    d=0
    sair=0
    pontos=0
    h=3
    pontos_desponiveis=1
    total=0
    book = openpyxl.load_workbook("questionarios.xlsx")
    #verifica se o teste existe
    while a != 1:
        time.sleep(1)
        os.system("cls")
        print("*******Modo Responder*******")
        x=0
        if sair == 1:
            print("\nDigite sair se quiser cancelar: ")
        resposta=input("\nDigite o nome do teste: ")
        if resposta.lower() =="sair" and sair == 1:
            return
            
        elif resposta in book.sheetnames:
            page=book[resposta]
            a=1
            book2 = openpyxl.load_workbook("respostas.xlsx")
            page2=book2[resposta]

            while x!=1:
                
                cell = page2.cell(row=b,column=1)
                cell6=page2.cell(row=1,column=1)
                #verifica se o criador está a tentar responder ao teste
                if cell6.value==email_log:
                    print("\nNão podes responder a um teste que criaste!")
                    time.sleep(1)
                    x=1
                    a=0
                    sair=1
                #verifica se uma conta com o mesmo email já existe
                elif cell.value == email_log:
                    print("\nEste teste já foi preenchido!")
                    time.sleep(1)
                    x=1
                    a=0
                    sair=1
                if cell.value != None:
                    b+=1
                else:
                    cell.value=email_log
                    x=1
                    print("\nEntrada foi um sucesso!")
                    time.sleep(1)
        else:
            print("\nTeste não existe!")
            sair=1
            time.sleep(1)

    cell2=page2.cell(row=b,column=2)
    cell3=page.cell(row=c,column=1)
    if cell3.value==None:
        d=1
    else:
        os.system("cls")
        print("\n***Perguntas de Verdadeiro ou Falso***")
    while d!=1:
        pontos_desponiveis=1
        h=3
        cell3=page.cell(row=c,column=1)
        cell4=page.cell(row=c,column=2)
        if cell3.value==None:
            d=1
            break
        print("\n",cell3.value)
        t = time.time()
        resposta= input("R: ")
        #verificaçao de resposta
        if resposta.lower() != "v" and resposta.lower() != "f" and resposta.lower() != "verdadeiro" and resposta.lower() != "falso":
            print("\nResposta invalida!")
        else:
            c+=1
            if resposta.lower() == "verdadeiro":
                resposta="v"
            if resposta.lower() == "falso":
                resposta="f"
            if resposta.lower() == cell4.value.lower():
                t2= time.time()
                t3 = t+h
                #desconta consuante o tempo que se demora a responder
                while t3<=t2:
                    pontos_desponiveis-=0.1
                    h+=1
                    t3 = t+h

                if pontos_desponiveis<0:
                    pontos_desponiveis=0
                pontos+=(round(pontos_desponiveis,1))
            total+=1

    c+=1
    d=0
    cell3=page.cell(row=c,column=1)
    if cell3.value==None:
        d=1
    else:
        os.system("cls")
        print("\n***Perguntas de Escolha Multipla***")
    while d!=1:
        pontos_desponiveis=1
        h=4
        f=3
        cell3=page.cell(row=c,column=1)
        cell4=page.cell(row=c,column=2)
        cell5=page.cell(row=c,column=f)
        if cell3.value==None:
            d=1
            break
        print("\n",cell3.value)
        while f<=6:
            cell5=page.cell(row=c,column=f)
            if f==3:
                g="A-"
            if f==4:
                g="B-"
            if f==5:
                g="C-"
            if f==6:
                g="D-"
            print(g,cell5.value,)
            f+=1
        t = time.time() 
        resposta= input("R: ")
        #verificaçao de resposta
        if resposta.lower() != "a" and resposta.lower() != "b" and resposta.lower() != "c" and resposta.lower() != "d":
            print("\nResposta invalida!")
        else:
            c+=1
            if resposta.lower() == cell4.value.lower():
                t2= time.time()
                t3 = t+h
                #desconta consuante o tempo que se demora a responder
                while t3<=t2:
                    pontos_desponiveis-=0.1
                    h+=1
                    t3 = t+h

                if pontos_desponiveis<0:
                    pontos_desponiveis=0
                pontos+=(round(pontos_desponiveis,1))
        total+=1
    cell2.value=pontos
    #resultado
    print("\n*****Resultado*****\n",cell2.value,"de",total)
    book2.save("respostas.xlsx")
    time.sleep(3)      