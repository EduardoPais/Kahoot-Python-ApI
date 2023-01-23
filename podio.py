import os
import time
import openpyxl
import msvcrt as m
reiniciar = '\033[0m'
vermelho = '\033[31m'
verde = '\033[32m'
azul = '\033[34m' 
def podio():
    sair=0
    a=0
    b=0
    c=2
    primeiro=0
    segundo=0
    terceiro=0
    segundo_nome="Não existe"
    terceiro_nome="Não existe"
    while a!=1:
        os.system("cls")
        if sair==1:
            print("\nDigite sair se quiser cancelar: ")
        resposta=input("\nDigite o nome do teste: ")
        if resposta.lower() == "sair":
            return
        book = openpyxl.load_workbook("respostas.xlsx")
        if resposta in book.sheetnames:
            page=book[resposta]
            a=1
        else:
            print(vermelho+"\nTeste não existe!"+reiniciar)
            sair=1
            time.sleep(1)
    while b!=1:
        cell=page.cell(row=c,column=2)
        cell2=page.cell(row=2,column=2)
        cell3=page.cell(row=c,column=1)
        if cell2.value==None:
            print(vermelho+"\nAinda ninguem respondeu a este teste!"+reiniciar)
            time.sleep(1)
            b=1
            return
        elif cell.value == None:
            b=1
        elif cell.value>primeiro:
            primeiro=cell.value
            primeiro_nome=cell3.value
        c+=1
    c=2
    b=0
    while b!=1:
        cell=page.cell(row=c,column=2)
        cell3=page.cell(row=c,column=1)
        if cell.value == None:
            b=1
        elif cell.value>segundo and cell3.value!= primeiro_nome:
            segundo=cell.value
            segundo_nome=cell3.value
        c+=1
    c=2
    b=0
    while b!=1:
        cell=page.cell(row=c,column=2)
        cell3=page.cell(row=c,column=1)
        if cell.value == None:
            b=1
        elif cell.value>terceiro  and cell3.value!= segundo_nome and cell3.value!= primeiro_nome:
            terceiro=cell.value
            terceiro_nome=cell3.value
        c+=1
    print(azul+"\n**********Podio**********"+reiniciar+"\n1º-",primeiro_nome,"-",primeiro,"\n2º-",segundo_nome,"-",segundo,"\n3º-",terceiro_nome,"-",terceiro)
    print("\nClique em qualquer tecla para continuar!")
    m.getch()

def lista_testes():
    os.system("cls")
    print(azul+"***********Testes***********"+reiniciar)
    book = openpyxl.load_workbook("questionarios.xlsx")
    print(book.sheetnames)
    print("\nClique em qualquer tecla para continuar!")
    m.getch()