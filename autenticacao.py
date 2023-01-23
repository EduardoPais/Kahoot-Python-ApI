import os  
import time
import openpyxl 
reiniciar = '\033[0m'
vermelho = '\033[31m'
verde = '\033[32m'
azul = '\033[34m' 

def signup():
    os.system("cls")
    print(azul+"\n************* Signup **************"+reiniciar)
    email_sig=str(input("Digite o seu email: "))    
    if email_sig =="":
        signup()
    last_char= email_sig[-1]
    firt_char= email_sig[0]
    n_a=0
    for char in email_sig:
        if char=="@":
            n_a+=1
        if char ==" ":
            signup()
    if last_char=="@" or firt_char=="@" or n_a!=1:
        signup()
    senha="a"
    senha2="b"
    a=0
    b=0
    c=0
    d=1
    while len(senha)>=13 or len(senha)<=7 or senha!=senha2 or c!=1 or a!=1 or b!=1 or d!=1:
        d=1
        a=0
        b=0
        c=0
        os.system("cls")
        print(azul+"\n************* Signup **************"+reiniciar)
        print("Didite o seu emial:",email_sig)
        print("\nPassword tem de conter entre 8 e 12 caracteres, conter no mínimo uma maiúscula, uma minúscula e um número.")
        senha=str(input("Digite a sua senha: "))
        senha2=str(input("Digite a sua senha novamente: "))
        for char in senha:
            if char.isnumeric():
                a=1
            elif char.islower():
                b=1
            elif char.isupper():
                c=1
            else:
                d=0

    n=1
    x=0
    book = openpyxl.load_workbook(r"contas.xlsx")
    sh = book.active
    while x!=1:
        cell = sh.cell(row=n,column=1)
        cell2 = sh.cell(row=n,column=2)
        if cell.value == email_sig:
            print(vermelho+"Usuario já existente!"+reiniciar)
            time.sleep(1)
            exec(open('main.py').read())

        if cell.value != None:
            n+=1
        else:
            cell.value=email_sig
            cell2.value=senha
            x=1
    book.save(r"contas.xlsx")
    print(verde+"\nO registo foi um sucesso."+reiniciar)
    time.sleep(1)

def login():
    x3=0
    sair=0
    while x3!=1:
        os.system("cls")
        print(azul+"\n************* Login **************"+reiniciar)
        n2=1
        x2=0
        global email_log
        if sair==1:
            print("Digite sair se quiser cancelar:\n")   
        email_log=str(input("Digite o seu email: "))
        if email_log.lower() =="sair" and sair == 1:
            exec(open('main.py').read())
        book = openpyxl.load_workbook(r"contas.xlsx")
        sh = book.active
        while x2!=1:
            cell = sh.cell(row=n2,column=1)
            cell2 = sh.cell(row=n2,column=2)
            if cell.value == email_log:
                x2=1
                x3=1  
            if cell.value == None:
                print(vermelho+"\nConta não existente!"+reiniciar)
                x2=1
                time.sleep(1)
                sair=1
            else:
                n2+=1
    n2-=1
    senha=str(input("Digite a sua senha: "))
    cell2 = sh.cell(row=n2,column=2)
    if cell2.value==senha:
        print(verde+"\nLogin com sucesso!"+reiniciar)
        time.sleep(1)
    else:
        print(vermelho+"\nLogin não foi possivel!"+reiniciar)
        time.sleep(1)
        login()
    book.save(r"contas.xlsx")